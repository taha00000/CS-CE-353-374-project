from openpyxl import load_workbook
from openpyxl import Workbook

# Load the Excel file
input_file = "your_file.xlsx"  # Replace with your file name
output_file = "transposed_file.xlsx"  # Output file name

# Load the workbook and select the active sheet
wb = load_workbook(input_file)
sheet = wb.active

# Create a new workbook for the output
output_wb = Workbook()
output_sheet = output_wb.active

# Transpose vertical blocks into horizontal rows
block_size = 10  # Number of rows in each vertical block
row_offset = 1  # Start writing from the first row in the output sheet

for start_row in range(1, sheet.max_row + 1, block_size):
    col_offset = 1  # Start writing from the first column in the output sheet
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + block_size - 1, min_col=1, max_col=1, values_only=True):
        for value in row:
            output_sheet.cell(row=row_offset, column=col_offset, value=value)
            col_offset += 1
    row_offset += 1  # Move to the next row for the next block

# Save the transposed data to a new file
output_wb.save(output_file)
print(f"Transposed data saved to {output_file}")