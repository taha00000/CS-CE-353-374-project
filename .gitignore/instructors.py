from openpyxl import load_workbook
from openpyxl import Workbook

# Load the Excel file
input_file = "transposed_file.xlsx"  # Replace with your file name
output_file = "data.xlsx"  # Output file name

# Load the workbook and select the active sheet
wb = load_workbook(input_file)
sheet = wb.active

# Create a new workbook for the output
output_wb = Workbook()
output_sheet = output_wb.active
row_offset = 2

day = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
names = ["program", "first_name", "last_name", "location", "day", "start_time", "end_time"]
for col in range(7):
    output_sheet.cell(row=1, column=col+1, value=names[col])

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9, values_only=True):
    break_while = False
    stopped = 4
    while not break_while:
        break_while = True
        for col in range(stopped, 9):
            if row[col] != "-":
                start_time, end_time = row[col].split("-")
                output_sheet.cell(row=row_offset, column=5, value=day[col - 4])

                time, am_pm = start_time.split()
                hour, min = time.split(":")
                if am_pm in ["PM", "pm", "pM", "Pm"]:
                    hour = int(hour) + 12
                output_sheet.cell(row=row_offset, column=6, value=str(hour)+":"+min)

                time, am_pm = end_time.split()
                hour, min = time.split(":")
                if am_pm in ["PM", "pm", "pM", "Pm"]:
                    hour = int(hour) + 12
                output_sheet.cell(row=row_offset, column=7, value=str(hour)+":"+min)

                for col_ in range(4):
                    output_sheet.cell(row=row_offset, column=col_+1, value=row[col_])
                row_offset += 1

                stopped = col+1
                break_while = False
                break
        

# Save the transposed data to a new file
output_wb.save(output_file)
print(f"Transposed data saved to {output_file}")